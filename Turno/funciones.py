import csv
import decimal
import os
from openpyxl import load_workbook
from pyexcel_ods3 import get_data

from main.models import Visitas, Turno, Cliente, Mascota

def registrar_visita_vacunacion(turno, datos):
    descripcion = datos.POST["descripcion"]
    
    cliente = Cliente.objects.get(pk=turno.cliente.pk)
    mascota = Mascota.objects.get(pk=turno.mascota.pk)
    codigo = datos.POST["codigo"]
    peso = datos.POST["peso"]
    monto= datos.POST["monto"]

    visita = Visitas(fecha=turno.fecha,
                            motivo=turno.motivo,
                            observaciones=descripcion,
                            cliente=cliente, 
                            mascota=mascota,
                            codigo=codigo,
                            peso=peso,
                            monto=monto)
    visita.save()
  


def registrar_visita_desparacitacion(turno, datos):
    descripcion = datos.POST["descripcion"]
    
    cliente = Cliente.objects.get(pk=turno.cliente.pk)
    mascota = Mascota.objects.get(pk=turno.mascota.pk)
    codigo = datos.POST["codigo"]
    peso = datos.POST["peso"]
    cantidad = datos.POST["cantidad"]
    monto = decimal.Decimal(datos.POST["monto"])
    #monto= datos.POST["monto"]

    visita = Visitas(fecha=turno.fecha,
                            motivo=turno.motivo,
                            observaciones=descripcion,
                            cliente=cliente, 
                            mascota=mascota,
                            codigo=codigo,
                            peso=peso,
                            cant_desparacitante=cantidad,
                            monto= monto)
    visita.save()
  

def registrar_visita_simple(descripcion, monto, turno):   
    cliente = Cliente.objects.get(pk=turno.cliente.pk)
    mascota = Mascota.objects.get(pk=turno.mascota.pk)

    visita = Visitas(fecha=turno.fecha,
                            motivo=turno.motivo,
                            observaciones=descripcion,
                            cliente=cliente, 
                            mascota=mascota,
                            monto=monto)
    visita.save()


def registrar_visita(turno, datos):

    descripcion = datos.POST["descripcion"]
    
    cliente = Cliente.objects.get(pk=turno.cliente.pk)
    mascota = Mascota.objects.get(pk=turno.mascota.pk)

    visita = Visitas.objects.create(fecha=turno.fecha,
                            motivo=turno.motivo,
                            observaciones=descripcion,
                            cliente=cliente, 
                            mascota=mascota)
    
    print("Registrando")

    if ("Vacunación" in turno.motivo):
        visita.peso = datos.POST["peso"]
        visita.codigo = datos.POST["codigo"]
    
    elif ("Desparacitación" in turno.motivo):
            print("Se registro la desparacitacion")
            visita.codigo = datos.POST["codigo"]
            visita.cant_desparacitante = datos.POST["cantidad"]
            visita.peso = datos.POST["peso"]
    print("Termino")

    visita.save()


def leer_archivo(archivo):
    csv_path = os.path.abspath(os.path.join(os.path.dirname(__file__), '../media/'+archivo))
    if archivo.endswith('.csv'):
        with open(csv_path) as file:
            reader = csv.reader(file)
            data = list(reader)
    elif archivo.endswith('.xlsx') or archivo.endswith('.lsx') :
        workbook = load_workbook(filename=archivo)
        sheet = workbook.active
        data = []
        for row in sheet.iter_rows(values_only=True):
            data.append(row)
    elif archivo.endswith('.ods'):
        arch = get_data(archivo)
        data = []
        
        for sheet_name in data:
            sheet_data = data[sheet_name]
            for row in sheet_data:
                for cell in row:
                    print(cell)
        
    else:
        data = []

    return data

def cacular_descuento(monto):
    return round(monto * 0.20 )
